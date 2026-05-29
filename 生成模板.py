import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ai_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
human_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fixed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
img_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # 浅红=图片列
header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
normal_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def make_sheet(wb, name, headers, col_types, sample_data, title):
    ws = wb.create_sheet(name)
    ncols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1,
        value="■ 红色=参考图(放文件名)  ■ 蓝色=AI自动填  ■ 黄色=人工填  ■ 绿色=固定配置")
    c.font = Font(name="微软雅黑", size=9, color="666666")

    for ci, (val, ct) in enumerate(zip(sample_data, col_types), 1):
        c = ws.cell(row=4, column=ci, value=val)
        c.font = normal_font; c.border = thin_border
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if ct == "图片": c.fill = img_fill
        elif ct == "AI": c.fill = ai_fill
        elif ct == "固定": c.fill = fixed_fill
        else: c.fill = human_fill

    for ri in range(5, 20):
        for ci in range(1, ncols + 1):
            c = ws.cell(row=ri, column=ci, value="")
            c.font = normal_font; c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ct = col_types[ci - 1]
            if ct == "图片": c.fill = img_fill
            elif ct == "AI": c.fill = ai_fill
            elif ct == "固定": c.fill = fixed_fill
            else: c.fill = PatternFill()

    # 列宽
    ws.column_dimensions["A"].width = 16  # 参考图
    ws.column_dimensions["B"].width = 14  # 款号
    ws.column_dimensions["C"].width = 34  # 标题
    for ci in range(4, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "A5"
    return ws

# ===== Sheet 1: 通用配置 =====
ws_cfg = wb.active
ws_cfg.title = "通用配置"
cfg_headers = ["配置项", "值", "说明"]
cfg_data = [
    ["品牌", "无品牌", ""],
    ["适用人群", "女性", ""],
    ["发货地", "广东省广州市", "改成实际发货地"],
    ["运费模板", "新疆西藏运费+10", ""],
    ["默认发货模式", "现货模式", "现货模式 / 现货+预售混合"],
    ["现货发货时间", "48小时内发货", "当日发/次日发/48h"],
    ["预售发货时间", "7天内发货", "3d/4d/5d/6d/7d/8d/9d/10d/15d"],
    ["七天无理由", "支持", ""],
    ["运费险", "支持", ""],
    ["商品状态", "下架", "审核后手动上架"],
    ["库存计数", "付款减库存", ""],
    ["面料材质", "聚酯纤维85% 其他15%", "固定填写"],
    ["参考图文件夹", r"商品图/参考图/", "AI从这里读图识别"],
    ["主图1x1文件夹", r"商品图/主图1x1/", "按款号命名"],
    ["主图3x4文件夹", r"商品图/主图3x4/", "按款号命名"],
]
ws_cfg.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
c = ws_cfg.cell(row=1, column=1, value="通用配置 — 全部商品共用，只填一次")
c.font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
c.alignment = Alignment(horizontal="center")
for ci, h in enumerate(cfg_headers, 1):
    c = ws_cfg.cell(row=2, column=ci, value=h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border
for i, rd in enumerate(cfg_data):
    for j, v in enumerate(rd):
        c = ws_cfg.cell(row=4 + i, column=j + 1, value=v)
        c.font = normal_font; c.fill = fixed_fill; c.border = thin_border
for i in range(len(cfg_data) + 4, 20):
    for j in range(1, 4):
        ws_cfg.cell(row=i, column=j).border = thin_border
ws_cfg.column_dimensions["A"].width = 18
ws_cfg.column_dimensions["B"].width = 24
ws_cfg.column_dimensions["C"].width = 42
ws_cfg.freeze_panes = "A4"

# ===== Sheet 2: 连衣裙 =====
make_sheet(wb, "连衣裙",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "领型", "袖长", "裙长", "裙型",
     "版型", "风格", "颜色", "材质", "图案", "腰型", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["001.jpg", "LK24001", "法式碎花收腰A字连衣裙女夏显瘦短裙", "V领", "短袖", "短裙", "A字型",
     "A字型", "法式", "白色", "雪纺", "碎花", "高腰", "夏季", "约会/日常",
     "89", "白色/黑色", "S/M/L", ""],
    "连衣裙 — 参考图放入 商品图\\参考图\\，1:1主图放入 商品图\\主图1x1\\，3:4主图放入 商品图\\主图3x4\\，均以款号命名")

# ===== Sheet 3: T恤-Polo衫 =====
make_sheet(wb, "T恤-Polo衫",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "领型", "袖长", "衣长", "版型",
     "风格", "颜色", "材质", "图案", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["002.jpg", "TX24001", "纯棉圆领短袖T恤女夏宽松百搭打底衫", "圆领", "短袖", "常规", "宽松",
     "休闲", "白色", "纯棉", "纯色", "夏季", "日常",
     "59", "白色/黑色", "S/M/L/XL", ""],
    "T恤")

# ===== Sheet 4: 蕾丝衫-雪纺衫 =====
make_sheet(wb, "蕾丝衫-雪纺衫",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "领型", "袖长", "衣长", "版型",
     "风格", "颜色", "材质", "图案", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["003.jpg", "LS24001", "法式蕾丝衫女夏方领泡泡袖短款显瘦上衣", "方领", "短袖", "短款", "修身",
     "法式", "白色", "蕾丝", "纯色", "夏季", "约会/日常",
     "79", "白色/黑色", "S/M/L", ""],
    "蕾丝衫")

# ===== Sheet 5: 针织衫-毛衣-羊毛衫 =====
make_sheet(wb, "针织衫-毛衣-羊毛衫",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "领型", "袖长", "衣长", "版型",
     "风格", "颜色", "材质", "图案", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["004.jpg", "ZZ24001", "韩版针织衫女春秋圆领长袖修身打底衫", "圆领", "长袖", "常规", "修身",
     "韩系", "米白色", "针织", "纯色", "春秋", "日常/通勤",
     "89", "米白色/黑色/灰色", "S/M/L", ""],
    "针织衫-毛衣-羊毛衫")

# ===== Sheet 6: 背心-吊带 =====
make_sheet(wb, "背心-吊带",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "领型", "衣长", "版型",
     "风格", "颜色", "材质", "图案", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["009.jpg", "BD24001", "韩版针织吊带背心女夏修身显瘦打底衫", "吊带", "短款", "修身",
     "韩系", "黑色", "针织", "纯色", "夏季", "日常/打底",
     "39", "黑色/白色/灰色", "S/M/L", ""],
    "背心-吊带 -- 无袖类，不需要袖长字段")

# ===== Sheet 7: 休闲裤 =====
make_sheet(wb, "休闲裤",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "裤长", "裤型", "腰型", "版型",
     "风格", "颜色", "材质", "图案", "厚薄", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["005.jpg", "XXK24001", "高腰阔腿休闲裤女春秋显瘦直筒长裤", "长裤", "阔腿", "高腰", "直筒",
     "韩系", "深蓝", "棉混纺", "纯色", "常规", "春秋", "日常/通勤",
     "129", "深蓝/黑色", "S/M/L/XL", ""],
    "休闲裤")

# ===== Sheet 8: 半身裙 =====
make_sheet(wb, "半身裙",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "裙长", "裙型", "腰型", "版型",
     "风格", "颜色", "材质", "图案", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["006.jpg", "BSQ24001", "韩系A字百褶半身裙女春秋高腰显瘦中长裙", "中长裙", "A字裙", "高腰", "A字型",
     "韩系", "黑色", "涤纶", "纯色", "春秋", "日常/约会",
     "79", "黑色", "S/M/L", ""],
    "半身裙")

# ===== Sheet 9: 外套 =====
make_sheet(wb, "外套",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字)", "领型", "袖长", "衣长", "版型",
     "风格", "颜色", "材质", "图案", "厚薄", "衣门襟", "适用季节", "适用场景",
     "售价\n(人工)", "规格颜色", "规格尺码", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工"],
    ["007.jpg", "WT24001", "韩版宽松西装外套女春秋通勤显瘦中长款", "西装领", "长袖", "中长款", "宽松",
     "韩系", "黑色", "涤纶", "纯色", "常规", "单排扣", "春秋", "通勤/日常",
     "159", "黑色/卡其色", "S/M/L", ""],
    "外套")

# ===== Sheet 10: 时尚套装 =====
make_sheet(wb, "时尚套装",
    ["参考图\n(填文件名)", "款号\n(人工)", "商品标题\n(AI ~30字 以套装结尾)", "上装领型", "上装袖长",
     "下装类型", "下装长度", "版型", "风格", "颜色", "材质", "图案", "适用季节", "适用场景",
     "售价\n(人工)", "上装尺码", "下装尺码", "颜色规格", "备注"],
    ["图片","人工","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","AI","人工","人工","人工","人工","人工"],
    ["008.jpg", "TZ24001", "韩系小香风外套+A字半裙时尚套装", "圆领", "长袖", "半身裙", "中长裙",
     "修身", "韩系", "米白色", "粗花呢", "纯色", "春秋", "通勤/约会",
     "199", "S/M/L", "S/M/L", "米白色", ""],
    '时尚套装 -- 标题必须以[套装]结尾，上下装尺码各填各的')

# 套装备注
ws9 = wb["时尚套装"]
note_row = 20
ws9.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=19)
n = ws9.cell(row=note_row, column=1, value="套装说明: 标题必须以[套装]结尾。上装尺码和下装尺码各填各的，相同直接填一样的。")
n.font = Font(name="微软雅黑", size=9, color="999999", italic=True)

# 每个Sheet底部加图片路径说明
for sn in ["连衣裙","T恤-Polo衫","蕾丝衫-雪纺衫","针织衫-毛衣-羊毛衫","背心-吊带","休闲裤","半身裙","外套","时尚套装"]:
    ws = wb[sn]
    note_row2 = 21
    ws.merge_cells(start_row=note_row2, start_column=1, end_row=note_row2, end_column=ws.max_column)
    n2 = ws.cell(row=note_row2, column=1,
        value=f"图片路径说明: 参考图→商品图\\参考图\\[文件名]  |  1:1主图→商品图\\主图1x1\\[款号].jpg  |  3:4主图→商品图\\主图3x4\\[款号].jpg")
    n2.font = Font(name="微软雅黑", size=9, color="999999", italic=True)

out = r"C:\Users\86153\Desktop\抖店上架助手\商品上架模板_v2.xlsx"
wb.save(out)
print("OK")
