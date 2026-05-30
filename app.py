"""
抖店上架助手 - Web 应用后端
"""
import json
import csv
import base64
import os
import re
from pathlib import Path
from io import BytesIO

from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
CORS(app)

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "商品图" / "参考图"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# 通义千问配置
DASHSCOPE_API_KEY = "sk-3cc39ec1fde843f48e049c943419d0ef"
DASHSCOPE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
MODEL_NAME = "qwen-vl-max"

# 加载下拉选项配置
OPTIONS_CONFIG = {}
try:
    with open(BASE_DIR / "下拉选项配置.json", "r", encoding="utf-8") as f:
        OPTIONS_CONFIG = json.load(f)
except:
    pass

# 加载字段配置表，获取每个类目的字段列表
def load_field_config():
    """读取字段配置表，返回 {类目: [字段列表]}"""
    config = {}
    try:
        with open(BASE_DIR / "字段配置表.csv", "r", encoding="gbk") as f:
            for row in csv.reader(f):
                if len(row) < 7 or not row[0].strip():
                    continue
                cat = row[0].strip()
                field = row[1].strip()
                if cat not in config:
                    config[cat] = []
                config[cat].append(field)
    except:
        pass
    return config


# ==================== 路由 ====================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config", methods=["GET"])
def get_config():
    """返回配置信息"""
    field_config = load_field_config()
    return jsonify({
        "options": OPTIONS_CONFIG,
        "fields": field_config,
        "categories": [k for k in field_config.keys() if k != "所有类目"]
    })


@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    """上传并解析 Excel 文件，同时提取A列嵌入的图片"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "未上传文件"}), 400

    # 保存文件
    path = BASE_DIR / "当前表格.xlsx"
    file.save(str(path))

    # 解析所有 Sheet
    wb = load_workbook(path)
    sheets_data = {}
    # 存图片的目录
    img_dir = BASE_DIR / "static" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)

    for sn in wb.sheetnames:
        if sn.startswith("_") or sn == "通用配置":
            continue
        ws = wb[sn]
        rows = []
        headers = []

        # 读取表头（第2行），跳过隐藏列
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=2, column=c).value
            if not h:
                headers.append(None)
                continue
            h = str(h).replace("\n", " ")
            col_letter = get_column_letter(c)
            col_dim = ws.column_dimensions.get(col_letter)
            # 检查列是否隐藏
            is_hidden = False
            if col_dim:
                try:
                    is_hidden = col_dim.hidden
                except:
                    pass
            if is_hidden:
                headers.append("__HIDDEN__")  # 标记为隐藏
            else:
                headers.append(h)

        # 提取A列的图片，按行号索引
        images_by_row = {}
        total_imgs = len(ws._images)
        for idx, img in enumerate(ws._images):
            row_num = None
            col_num = 1  # 默认A列
            try:
                if hasattr(img.anchor, '_from'):
                    row_num = img.anchor._from.row + 1
                    col_num = img.anchor._from.col + 1
            except:
                pass
            # 只要有行号且在数据区就提取
            if row_num and row_num >= 5:
                if row_num not in images_by_row:
                    images_by_row[row_num] = []
                img_filename = f"{sn}_r{row_num}_i{idx}.png"
                img_path = img_dir / img_filename
                try:
                    img_data = img._data()
                    # 跳过空数据
                    if not img_data or len(img_data) < 100:
                        continue
                    with open(img_path, "wb") as f:
                        f.write(img_data)
                    images_by_row[row_num].append(f"/static/images/{img_filename}")
                except Exception as ex:
                    print(f"图片提取失败 {sn} row={row_num}: {ex}")
                    pass

        # 过滤出可见表头
        visible_headers = [h for h in headers if h is not None and h != "__HIDDEN__"]

        # 找到最后一行有效数据，跳过底部的说明行（如SKU说明）
        last_data_row = 5
        for r in range(5, ws.max_row + 1):
            has_content = False
            for c in range(1, len(headers) + 1):
                if headers[c-1] is None or headers[c-1] == "__HIDDEN__":
                    continue
                val = ws.cell(row=r, column=c).value
                if val is not None and str(val).strip():
                    has_content = True
                    break
            # 如果这行第一个非隐藏列有内容且包含"说明"且不是商品数据，跳过
            first_col_val = ""
            for c in range(1, len(headers) + 1):
                h = headers[c-1]
                if h and h != "__HIDDEN__":
                    v = ws.cell(row=r, column=c).value
                    first_col_val = str(v).strip() if v else ""
                    break
            # 跳过底部说明行
            if first_col_val and ("说明" in first_col_val or "图片路径" in first_col_val):
                continue
            if has_content:
                last_data_row = r

        # 读取数据行（第5行起到底部说明行之前）
        for r in range(5, last_data_row + 1):
            row_data = {}
            non_empty_count = 0
            for c in range(1, len(headers) + 1):
                h = headers[c - 1]
                if h is None or h == "__HIDDEN__":
                    continue
                val = ws.cell(row=r, column=c).value
                if val is not None and str(val).strip():
                    row_data[h] = str(val).strip()
                    non_empty_count += 1
                else:
                    row_data[h] = ""

            # 添加A列的图片
            if r in images_by_row:
                row_data["_images"] = images_by_row[r]
                non_empty_count += 1

            # A列有图片文件名也算有参考图
            ref_has_image = False
            for hk, hv in row_data.items():
                if "参考图" in hk and hv and "." in hv:
                    ref_has_image = True
                    break

            # 任一列有实际内容就算有效行
            has_image = r in images_by_row or ref_has_image
            if has_image or non_empty_count >= 1:
                row_data["_row"] = r
                row_data["_category"] = sn
                rows.append(row_data)

        # 最后一行有效数据后，追加10个空行供编辑（标记为_placeholder）
        last_valid_row = rows[-1]["_row"] if rows else 4
        for r in range(last_valid_row + 1, last_valid_row + 11):
            row_data = {"_row": r, "_category": sn, "_placeholder": True}
            for c in range(1, len(headers) + 1):
                h = headers[c - 1]
                if h is None or h == "__HIDDEN__":
                    continue
                row_data[h] = ""
            rows.append(row_data)

        sheets_data[sn] = {
            "headers": visible_headers,
            "rows": rows
        }

    wb.close()
    return jsonify({"sheets": sheets_data, "sheetNames": list(sheets_data.keys())})


@app.route("/api/upload-image", methods=["POST"])
def upload_image():
    """上传参考图，返回可访问的路径"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "未上传文件"}), 400

    # 保存到static/images/ 确保可访问
    img_dir = BASE_DIR / "static" / "images"
    img_dir.mkdir(parents=True, exist_ok=True)

    import time
    # 清理文件名中的特殊字符（#会导致URL截断）
    safe_name = file.filename.replace('#','-').replace(' ','_')
    filename = f"{int(time.time()*1000)}_{safe_name}"
    file.save(str(img_dir / filename))

    # 返回相对路径
    img_url = f"/static/images/{filename}"
    return jsonify({"ok": True, "filename": filename, "url": img_url, "path": img_url})


@app.route("/api/ai-recognize", methods=["POST"])
def ai_recognize():
    """调用通义千问识图，识别服装属性"""
    data = request.json
    image_path = data.get("image_path", "")  # /static/images/xxx.png
    image_base64 = data.get("image_base64", "")  # 直接传base64
    category = data.get("category", "")
    fields_to_recognize = data.get("fields", [])
    sku_hint = data.get("sku_hint", "")  # SKU文本提示（如"黑色连衣裙"）

    # 获取图片base64
    if image_base64:
        image_data = image_base64
    elif image_path:
        # 尝试多种路径
        full_path = None
        # 1. static/images/ 路径
        p = BASE_DIR / image_path.lstrip("/")
        if p.exists():
            full_path = p
        # 2. 商品图/参考图/ 目录下查找
        if not full_path:
            ref_dir = BASE_DIR / "商品图" / "参考图"
            for ext in ['.jpg','.jpeg','.png','.webp','.bmp']:
                candidate = ref_dir / (image_path.replace('/static/images/','').split('/')[-1])
                if candidate.exists():
                    full_path = candidate
                    break
                # 尝试原始文件名
                candidate2 = ref_dir / image_path
                if candidate2.exists():
                    full_path = candidate2
                    break
        if not full_path:
            return jsonify({"error": f"图片不存在: {image_path}"}), 404
        with open(full_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode()
    else:
        return jsonify({"error": "未提供图片"}), 400

    # 获取该类目的下拉选项
    options = OPTIONS_CONFIG.get(category, {})
    all_category_options = OPTIONS_CONFIG.get("所有类目", {})
    merged_options = {**all_category_options, **options}

    # 构建提示词
    fields_desc = []
    for fn in fields_to_recognize:
        opts = merged_options.get(fn, {}).get("options", "")
        if not opts:
            opts = merged_options.get(fn, {}).get("fixed", "")
        hint = f"- {fn}"
        if opts:
            opts_list = [o.strip() for o in opts.replace("/", ",").split(",") if o.strip()]
            if opts_list:
                hint += f"（从以下选一个最接近的：{'/'.join(opts_list[:30])}）"
        fields_desc.append(hint)

    sku_context = ""
    if sku_hint:
        sku_context = f"""\n\n【SKU信息 - 人工标注，必须优先采用】: {sku_hint}

识别原则：SKU信息 > 图片识别。先看SKU判断品类，再看图片识别具体属性。

对于时尚套装类目：
- SKU有"裙子"或"半身裙" → 下装类型=半身裙，下装长度根据图片选(短裙/中长裙/长裙)
- SKU有"裤子"或"休闲裤" → 下装类型=休闲裤，下装长度根据图片选
- SKU有"上衣"或"外套" → 上装类型据此判断
- SKU有"套装" → 确认是套装商品，一个SKU对应上衣+下装的组合

结合SKU和图片，综合判断所有属性。SKU说是什么品类就是什么品类，图片只看具体款式细节。"""

    prompt = f"""请识别这张服装图片的属性。这是一个{category}类目的商品。{sku_context}

请返回以下属性(JSON格式):
{chr(10).join(fields_desc)}

注意：
1. 只返回JSON，不要其他文字
2. 每个字段必须从给出的选项中选最接近的一个值，不能自己编
3. 如果图片中无法判断某个属性，填""
4. 颜色有多个时用/分隔如"白色/黑色"
5. 不要用模糊描述如"时尚""好看"，从选项里选具体的
6. JSON的key用中文，与上面字段名完全一致

返回格式: {{"领型": "圆领", "袖长": "短袖", "颜色": "白色"}}"""

    try:
        resp = requests.post(
            DASHSCOPE_URL,
            headers={"Authorization": f"Bearer {DASHSCOPE_API_KEY}"},
            json={
                "model": MODEL_NAME,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}},
                        {"type": "text", "text": prompt}
                    ]
                }],
                "max_tokens": 500
            },
            timeout=30
        )
        result = resp.json()

        if "choices" not in result:
            return jsonify({"error": f"API返回异常: {json.dumps(result, ensure_ascii=False)[:200]}"}), 500

        content = result["choices"][0]["message"]["content"]

        # 提取JSON
        json_match = re.search(r'\{[^}]+\}', content, re.DOTALL)
        attrs = {}
        if json_match:
            attrs = json.loads(json_match.group())

        # 自动生成标题
        season = attrs.get("适用季节", "夏季")
        title = generate_title_for_category(category, attrs, season)

        return jsonify({"ok": True, "attributes": attrs, "title": title.get("title", ""),
                        "short_title": title.get("short_title", ""), "raw": content})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def generate_title_for_category(category, attributes, season="夏季"):
    """根据识别的属性生成标题"""
    attr_desc = " ".join([f"{k}:{v}" for k, v in attributes.items() if v])
    if not attr_desc:
        return {}

    prompt = f"""你是抖音服装电商运营专家。请模仿抖音爆款服装链接的标题风格生成标题。

类目：{category}
属性：{attr_desc}
季节：{season}

【严格规则 - 违反即错误】
1. 标题30字，短标题12字
2. 类目是{category}，标题品类词必须与此一致
3. 仔细看属性信息：如果属性里写了"半身裙"或"裙子"，标题绝对不能出现"裤子""休闲裤"
4. 如果属性里写了"休闲裤"或"裤子"，标题绝对不能出现"裙子""半身裙"
5. 标题以类目关键词为准，类目是连衣裙就写连衣裙，是套装就以"套装"结尾

爆款标题风格要求：
- 关键词密集，把属性词紧凑排列，不浪费任何一个字
- 不要写材质面料（如纯棉/雪纺/涤纶等）
- 不要出现"单件""一件""单条"等词
- 不要出现年龄范围如"18-29岁"、不要说"穿搭""搭配""清凉"等泛词
- 不用"适合""可以""非常"等废话
- 类目关键词必须出现（连衣裙/T恤/休闲裤等）
- 风格词+品类词+属性词+季节+卖点词 紧凑组合
- 参考风格：像"法式复古碎花收腰显瘦A字连衣裙女夏中长裙"这种

规则：标题必须30字，短标题12字通顺完整，年份2026，套装标题以"套装"结尾。
只返回JSON：{{"title":"30字标题","short_title":"12字短标题"}}"""

    try:
        resp = requests.post(
            DASHSCOPE_URL,
            headers={"Authorization": f"Bearer {DASHSCOPE_API_KEY}"},
            json={"model": "qwen-plus", "messages": [{"role": "user", "content": prompt}], "max_tokens": 200},
            timeout=15
        )
        content = resp.json()["choices"][0]["message"]["content"]
        m = re.search(r'\{[^}]+\}', content, re.DOTALL)
        if m:
            return json.loads(m.group())
    except:
        pass
    return {}


@app.route("/api/generate-title", methods=["POST"])
def generate_title():
    """生成30字商品标题和12字短标题"""
    data = request.json
    category = data.get("category", "")
    attributes = data.get("attributes", {})
    season = data.get("season", "夏季")

    # 构建属性描述
    attr_desc = " ".join([f"{k}:{v}" for k, v in attributes.items() if v])

    prompt = f"""你是一个抖音服装电商运营专家。请根据以下信息生成商品标题。

类目：{category}
属性：{attr_desc}
季节：{season}

要求：
1. 标题正好30个字，模仿抖音爆款风格
2. 标题必须包含类目关键词（如"连衣裙""T恤"等）
3. 如果是时尚套装，标题必须以"套装"结尾
4. 年份使用2026
5. 同时生成一个12个字的短标题（完整通顺的摘要，不能是截断的）
6. 只返回JSON：{{"title": "30字标题", "short_title": "12字短标题"}}
7. 不要任何其他文字"""

    try:
        resp = requests.post(
            DASHSCOPE_URL,
            headers={"Authorization": f"Bearer {DASHSCOPE_API_KEY}"},
            json={
                "model": "qwen-plus",
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 200
            },
            timeout=20
        )
        result = resp.json()
        content = result["choices"][0]["message"]["content"]
        json_match = re.search(r'\{[^}]+\}', content, re.DOTALL)
        if json_match:
            title_data = json.loads(json_match.group())
            return jsonify({"ok": True, **title_data})
        return jsonify({"ok": False, "raw": content})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/save-excel", methods=["POST"])
def save_excel():
    """保存编辑后的数据到 Excel"""
    data = request.json
    sheets_data = data.get("sheets", {})

    path = BASE_DIR / "当前表格.xlsx"
    wb = load_workbook(path)

    for sn, sheet_info in sheets_data.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        headers = sheet_info.get("headers", [])
        rows = sheet_info.get("rows", [])

        # 构建表头到列号的映射
        header_to_col = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=2, column=c).value
            if h:
                header_to_col[str(h).replace("\n", " ").strip()] = c

        for row_data in rows:
            row_num = row_data.get("_row", 0)
            if not row_num:
                continue
            for h, val in row_data.items():
                if h.startswith("_"):
                    continue
                col_num = header_to_col.get(h)
                if col_num and val:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = val

    wb.save(str(path))
    wb.close()
    return jsonify({"ok": True})


@app.route("/api/export-excel", methods=["GET"])
def export_excel():
    """下载当前 Excel"""
    path = BASE_DIR / "当前表格.xlsx"
    if path.exists():
        return send_file(str(path), as_attachment=True, download_name="商品上架模板.xlsx")
    return jsonify({"error": "文件不存在"}), 404


@app.route("/api/output-tables", methods=["GET"])
def list_output_tables():
    """列出已完成的表格"""
    output_dir = BASE_DIR / "已完成的表格"
    output_dir.mkdir(parents=True, exist_ok=True)
    tables = []
    for f in sorted(output_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        tables.append({
            "name": f.name,
            "size": f"{f.stat().st_size / 1024:.1f} KB",
            "time": f.stat().st_mtime
        })
    return jsonify({"tables": tables})


@app.route("/api/save-output", methods=["POST"])
def save_output():
    """确认表格完成，保存到输出目录"""
    data = request.json
    table_name = data.get("name", "").strip()
    if not table_name:
        return jsonify({"error": "请输入表格名称"}), 400

    # 保存编辑后的数据
    sheets_data = data.get("sheets", {})
    path = BASE_DIR / "当前表格.xlsx"
    if path.exists():
        wb = load_workbook(path)
        for sn, sheet_info in sheets_data.items():
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            headers = sheet_info.get("headers", [])
            rows = sheet_info.get("rows", [])
            header_to_col = {}
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=2, column=c).value
                if h:
                    header_to_col[str(h).replace("\n", " ").strip()] = c
            for row_data in rows:
                row_num = row_data.get("_row", 0)
                if not row_num:
                    continue
                for h, val in row_data.items():
                    if h.startswith("_"):
                        continue
                    col_num = header_to_col.get(h)
                    if col_num and val:
                        ws.cell(row=row_num, column=col_num).value = val
        wb.save(str(path))
        wb.close()

    # 复制到输出目录
    import shutil
    output_dir = BASE_DIR / "已完成的表格"
    output_dir.mkdir(parents=True, exist_ok=True)
    dest = output_dir / f"{table_name}.xlsx"
    shutil.copy(str(path), str(dest))

    return jsonify({"ok": True, "name": f"{table_name}.xlsx"})


@app.route("/api/delete-output", methods=["POST"])
def delete_output():
    """删除已完成的表格"""
    data = request.json
    name = data.get("name", "")
    path = BASE_DIR / "已完成的表格" / name
    if path.exists():
        path.unlink()
        return jsonify({"ok": True})
    return jsonify({"error": "文件不存在"}), 404


@app.route("/api/download-output/<name>", methods=["GET"])
def download_output(name):
    """下载已完成的表格"""
    path = BASE_DIR / "已完成的表格" / name
    if path.exists():
        return send_file(str(path), as_attachment=True, download_name=name)
    return jsonify({"error": "文件不存在"}), 404


@app.route("/api/start-auto-upload", methods=["POST"])
def start_auto_upload():
    """启动浏览器自动化上架"""
    data = request.json
    sheet_name = data.get("sheet", "")
    row_index = data.get("rowIndex", 0)

    if not sheet_name:
        return jsonify({"error": "请指定Sheet"}), 400

    # 读取当前表格数据
    path = BASE_DIR / "当前表格.xlsx"
    if not path.exists():
        return jsonify({"error": "请先上传Excel"}), 400

    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return jsonify({"error": f"Sheet不存在: {sheet_name}"}), 400

    ws = wb[sheet_name]

    # 读表头
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=2, column=c).value
        if h:
            headers[str(h).replace("\n", " ").strip()] = c

    # 读指定行数据
    row_num = row_index + 5  # 数据从第5行开始
    row_data = {}
    for h, c in headers.items():
        val = ws.cell(row=row_num, column=c).value
        row_data[h] = str(val).strip() if val else ""

    wb.close()

    # 启动 Playwright 自动化
    try:
        result = run_browser_automation(sheet_name, row_data, headers)
        return jsonify({"ok": True, "message": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def run_browser_automation(category, row_data, headers):
    """浏览器自动上架：上传图片+填属性+填价格库存"""
    from playwright.sync_api import sync_playwright

    NL = chr(10)
    messages = []
    style_code = row_data.get("款号", "").strip()

    # 收集标题
    title = ""
    for k, v in row_data.items():
        if "商品标题" in k and v:
            title = v
            break
    if not title:
        return "错误：没有商品标题"

    # 收集所有要填的属性
    attrs = {}
    for k, v in row_data.items():
        if v and k not in ["_row", "_category", "_images", "_placeholder"] and "参考图" not in k and "SKU" not in k and "售价" not in k and "款号" not in k:
            attrs[k] = v

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=str(BASE_DIR / "browser_data"),
            headless=False, args=["--disable-blink-features=AutomationControlled"]
        )
        page = ctx.new_page()
        # 进入抖店主页
        page.goto("https://fxg.jinritemai.com/")
        page.wait_for_timeout(3000)
        messages.append("已打开抖店主页")

        # 点击左侧「创建商品」
        try:
            page.locator('text=创建商品').first.click()
            page.wait_for_timeout(3000)
            messages.append("已进入商品发布页")
        except:
            # 备用：直接访问创建页面
            page.goto("https://fxg.jinritemai.com/ffa/merchant/product/create")
            page.wait_for_timeout(3000)
            messages.append("备用方式进入发布页")

        # 登录检查
        if "login" in page.url.lower() or "passport" in page.url.lower():
            messages.append("请扫码登录...")
            try:
                page.wait_for_url("**/product/**", timeout=300000)
            except:
                return "登录超时"

        # ---- 上传1:1主图 ----
        img_dir = BASE_DIR / "商品图" / "主图1x1"
        imgs = sorted(list(img_dir.glob(f"{style_code}*"))) if style_code else []
        if imgs:
            try:
                page.locator('input[type="file"]').first.set_input_files([str(p) for p in imgs[:5]])
                messages.append(f"1:1主图 {len(imgs[:5])}张")
            except Exception as e:
                messages.append(f"主图上传失败:{e}")
        page.wait_for_timeout(3000)

        # ---- 填标题 ----
        try:
            page.locator('input[placeholder*="标题"]').first.fill(title)
            messages.append("标题已填")
            page.wait_for_timeout(2000)
        except:
            messages.append("标题填写失败")

        # ---- 上传3:4主图 ----
        imgs34 = sorted(list((BASE_DIR / "商品图" / "主图3x4").glob(f"{style_code}*"))) if style_code else []
        if imgs34:
            try:
                fis = page.locator('input[type="file"]')
                if fis.count() > 1:
                    fis.nth(1).set_input_files([str(p) for p in imgs34[:5]])
                    messages.append(f"3:4主图 {len(imgs34[:5])}张")
            except:
                pass

        # ---- 上传视频 ----
        vids = sorted(list((BASE_DIR / "商品图" / "主图视频").glob(f"{style_code}*"))) if style_code else []
        if vids:
            try:
                fis = page.locator('input[type="file"]')
                if fis.count() > 2:
                    fis.nth(2).set_input_files([str(vids[0])])
                    messages.append("视频已上传")
            except:
                pass
        page.wait_for_timeout(2000)

        # ---- 点下一步 ----
        try:
            page.locator('button:has-text("下一步")').first.click()
            messages.append("进入详情页")
            page.wait_for_timeout(3000)
        except:
            messages.append("请手动点下一步")

        # ---- 填属性（品牌/材质等） ----
        for field, val in attrs.items():
            if not val: continue
            try:
                el = page.locator(f'div:has-text("{field}")').first
                if el.is_visible():
                    el.click()
                    page.wait_for_timeout(300)
                    opt = page.locator(f'text="{val}"').first
                    if opt.is_visible():
                        opt.click()
                        messages.append(f"{field}={val}")
            except:
                pass

        # ---- 发布（选下架） ----
        try:
            # 找到发布/提交按钮
            submit_btn = page.locator('button:has-text("发布")').first
            if not submit_btn.is_visible(timeout=2000):
                submit_btn = page.locator('button:has-text("提交")').first
            if submit_btn.is_visible():
                # 先确保选了下架
                offline_radio = page.locator('text=下架').first
                if offline_radio.is_visible():
                    offline_radio.click()
                    page.wait_for_timeout(500)
                    messages.append("已选下架")
                submit_btn.click()
                page.wait_for_timeout(3000)
                messages.append("已提交（下架状态），请人工审核后上架")
        except:
            messages.append("请手动选择下架并点击发布")

        page.screenshot(path=str(BASE_DIR / "debug_screenshot.png"))
        page.wait_for_timeout(10000)

    return NL.join(messages)
if __name__ == "__main__":
    print("抖店上架助手启动中...")
    print("打开浏览器访问: http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
