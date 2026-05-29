import csv, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ===== 1. 读取字段配置表 (GBK) =====
all_rows = []
with open(r'C:\Users\86153\Desktop\抖店上架助手\字段配置表.csv', 'r', encoding='gbk') as f:
    for row in csv.reader(f):
        all_rows.append(row)

data_rows = all_rows[1:]

categories = {}
cat_order = []
for row in data_rows:
    if len(row) < 7:
        continue
    cat = row[0].strip()
    if not cat:
        continue
    field_info = {
        'field': row[1].strip(),
        'type': row[2].strip() if len(row) > 2 else '',
        'required': row[3].strip() if len(row) > 3 else '',
        'ai': row[4].strip() if len(row) > 4 else '',
        'editable': row[5].strip() if len(row) > 5 else '',
        'options': row[6].strip() if len(row) > 6 else '',
        'fixed': row[7].strip() if len(row) > 7 else '',
        'note': row[8].strip() if len(row) > 8 else '',
    }
    if cat not in categories:
        categories[cat] = []
        cat_order.append(cat)
    categories[cat].append(field_info)

common_fields = categories.get('所有类目', [])

# Sheet顺序定义
sheet_order = ['时尚套装', '连衣裙', 'T恤-Polo衫-衬衫', '蕾丝衫-雪纺衫',
               '针织衫-毛衣-羊毛衫', '背心-吊带', '休闲裤', '半身裙', '外套']

print("=== 类目及字段数量 ===")
for cat in [c for c in cat_order if c != '所有类目']:
    filled = sum(1 for f in categories[cat] if f['options'] or f['fixed'])
    print(f"  {cat}: {len(categories[cat])}字段, {filled}已填")

# ===== 2. 重建 类目属性填写表.csv =====
with open(r'C:\Users\86153\Desktop\抖店上架助手\类目属性填写表.csv', 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['类目', '属性字段', '抖店显示名', '字段类型', '可选值(请填写/逗号分隔)', '填写示例'])
    for cat in sheet_order:
        if cat in categories:
            for fi in categories[cat]:
                row = [cat, fi['field'], '', fi['type'], fi['options'], '']
                writer.writerow(row)
print("已重建: 类目属性填写表.csv")

# ===== 3. 生成 Excel 模板 =====
wb = openpyxl.Workbook()

hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
aifill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
hufill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fxfill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
imfill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
hfont = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
nfont = Font(name="微软雅黑", size=10)
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

# --- Sheet 1: 通用配置 ---
ws = wb.active
ws.title = "通用配置"
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
c = ws.cell(row=1, column=1, value="通用配置 - 全部商品共用，设一次即可")
c.font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
c.alignment = Alignment(horizontal="center")

for ci, h in enumerate(["配置项", "值", "说明"], 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = hfont; c.fill = hfill; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")

cfg_items = []
for fi in common_fields:
    val = fi['fixed'] or fi['options']
    cfg_items.append([fi['field'], val, fi['note']])

extra_cfg = [
    ["面料材质", "聚酯纤维85% 其他15%", "固定填写"],
    ["商品状态", "下架（审核后手动上架）", ""],
    ["库存计数", "付款减库存", ""],
    ["参考图文件夹", "商品图/参考图/", "AI识别用"],
    ["主图1x1文件夹", "商品图/主图1x1/", "按款号命名"],
    ["主图3x4文件夹", "商品图/主图3x4/", "按款号命名"],
]
cfg_items.extend(extra_cfg)

for i, rd in enumerate(cfg_items):
    for j, v in enumerate(rd):
        c = ws.cell(row=4 + i, column=j + 1, value=v)
        c.font = nfont; c.fill = fxfill; c.border = border
for i in range(len(cfg_items) + 4, 22):
    for j in range(1, 4):
        ws.cell(row=i, column=j).border = border
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 42
ws.freeze_panes = "A4"

# --- 每个类目一个 Sheet ---
# 30字抖音风格标题，必须含类目关键词
sample_titles = {
    '时尚套装': '2026春款韩系小香风外套半身裙时尚套装女通勤两件套装',
    '连衣裙': '2026夏款法式复古碎花收腰A字连衣裙女显瘦中长裙子',
    'T恤-Polo衫-衬衫': '2026夏款纯棉圆领短袖T恤女宽松百搭打底上衣夏装',
    '蕾丝衫-雪纺衫': '2026夏款法式方领泡泡袖蕾丝衫女短款显瘦雪纺上衣',
    '针织衫-毛衣-羊毛衫': '2026春秋韩版圆领长袖针织衫女修身打底毛衣羊毛衫',
    '背心-吊带': '2026夏款韩版针织吊带背心女修身显瘦打底外穿上衣',
    '休闲裤': '2026春秋款高腰阔腿休闲裤女显瘦直筒长裤宽松垂感',
    '半身裙': '2026春秋款韩系A字百褶半身裙女高腰显瘦中长裙子',
    '外套': '2026春秋款韩版宽松西装外套女通勤显瘦中长款上衣',
}

# 12字短标题（完整通顺的摘要，不能截断）
short_titles = {
    '时尚套装': '小香风外套半身裙时尚套装',
    '连衣裙': '法式碎花收腰显瘦连衣裙',
    'T恤-Polo衫-衬衫': '纯棉圆领短袖T恤打底衫',
    '蕾丝衫-雪纺衫': '法式方领泡泡袖蕾丝衫女装',
    '针织衫-毛衣-羊毛衫': '韩版圆领长袖针织打底衫',
    '背心-吊带': '韩版针织吊带背心打底衫',
    '休闲裤': '高腰阔腿显瘦直筒休闲裤',
    '半身裙': '韩系A字百褶显瘦半身裙',
    '外套': '韩版宽松西装通勤女装外套',
}

# 标题规范说明
title_rule = "标题=30字，须含类目关键词，年份按上市季节"
short_title_rule = "短标题=12字"

for cat in sheet_order:
    if cat not in categories:
        continue

    # 合并公共字段（所有类目中非全局配置的字段）
    # 这些字段已由硬编码列覆盖，不重复合并
    skip_merge = ['品牌', '运费模板', '七天无理由', '运费险', '发货地',
                  '售后服务承诺', '商品标题', '类目', '短标题',
                  '规格颜色', '规格尺码', 'SKU价格', 'SKU库存', '售价',
                  '主图(1:1)', '详情图(3:4)', '面料材质',
                  '规格颜色名', '颜色值', '规格尺码名', '尺码值']
    merged_common = []
    cat_field_names = {f['field'] for f in categories[cat]}
    for fi in common_fields:
        fn = fi['field']
        if fn in skip_merge or fn in cat_field_names:
            continue
        merged_common.append(fi)

    fields = merged_common + categories[cat]

    # 分三组: 人工 → AI → 固定
    manual_headers = []
    manual_types = []
    ai_headers = []
    ai_types = []
    fixed_headers = []
    fixed_types = []

    # 人工区: 参考图 + 款号 + SKU + 售价
    manual_headers = [
        "参考图\n(填文件名)",
        "款号\n(人工)",
        "SKU1\n(颜色)", "售价1\n(元)",
        "SKU2\n(颜色)", "售价2\n(元)",
        "SKU3\n(颜色)", "售价3\n(元)",
        "SKU4\n(颜色)", "售价4\n(元)",
        "SKU5\n(颜色)", "售价5\n(元)",
        "SKU6\n(颜色)", "售价6\n(元)",
        "SKU7\n(颜色)", "售价7\n(元)",
        "SKU8\n(颜色)", "售价8\n(元)",
    ]
    manual_types = (["图片", "人工"] +
                    ["人工", "人工"] * 8)  # 8组SKU+售价

    # 从公共/类目字段中分拣人工和固定
    for fi in fields:
        if fi['ai'] == '是':
            continue  # AI字段后面统一处理
        if fi['fixed'] and not fi['options']:
            fixed_headers.append(fi['field'])
            fixed_types.append("固定")
        else:
            manual_headers.append(fi['field'])
            manual_types.append("人工")

    # AI区: 标题 + 短标题 + 公共AI字段 + 类目AI字段
    ai_headers = ["商品标题\n(AI 30字)", "短标题\n(AI 12字)"]
    ai_types = ["AI", "AI"]
    for fi in fields:
        if fi['ai'] == '是':
            ai_headers.append(fi['field'])
            ai_types.append("AI")

    # 固定区头（规格尺码已去掉，尺码大小从字段配置来）

    # 备注放最后
    fixed_headers.append("备注")
    fixed_types.append("人工")

    # 合并
    headers = manual_headers + ai_headers + fixed_headers
    col_types = manual_types + ai_types + fixed_types
    fields = fields  # 保持fields引用用于后续下拉验证

    ncols = len(headers)
    sn = cat[:31]

    ws = wb.create_sheet(sn)

    # 标题行
    title_text = f"{cat} - 红色=参考图 | 黄色=人工填写 | 绿色=固定 | 蓝色=AI生成(有下拉)"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    # 表头
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 图例 + 规则
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1,
        value=f"红色=参考图 | 黄色=人工填 | 绿色=固定 | 蓝色=AI填(有下拉) | {title_rule} | {short_title_rule}")
    c.font = Font(name="微软雅黑", size=9, color="666666")

    # 构建示例行 (按headers顺序)
    def build_sample(field_name, fi=None):
        """根据字段名和配置生成示例值"""
        if field_name.startswith('参考图'): return "001.jpg"
        if field_name.startswith('款号'): return cat[:4] + "24001"
        if field_name.startswith('SKU1'): return "黑色"
        if field_name.startswith('售价1'): return "99"
        if field_name.startswith('SKU2'): return "白色"
        if field_name.startswith('售价2'): return "99"
        if 'SKU' in field_name and field_name.replace('SKU','').split('(')[0].isdigit():
            return ""
        if '售价' in field_name and field_name.replace('售价','').split('(')[0].isdigit():
            return ""
        if field_name.startswith('商品标题'): return sample_titles.get(cat, "商品标题示例")
        if field_name.startswith('短标题'):
            return short_titles.get(cat, "短标题示例")
        if field_name.startswith('备注'): return ""
        if field_name.startswith('尺码大小'): return "S:80-95斤 M:95-105斤 L:105-115斤 XL:115-125斤"
        if fi:
            opts = fi['options'] or fi['fixed']
            if opts and '/' in opts:
                return opts.split('/')[0].strip()
            if fi['options']:
                return fi['options'].replace('/', ',').split(',')[0].strip()
            if fi['fixed']:
                return fi['fixed']
        return " "

    # 构建字段名到fi的映射
    field_map = {}
    for fi in fields:
        field_map[fi['field']] = fi

    sample = []
    for h in headers:
        # 去掉换行和标记提取纯字段名
        fn = h.split('\n')[0].replace('*', '').strip()
        fi = field_map.get(fn, None)
        sample.append(build_sample(fn, fi))

    # 写示例行和数据行
    for ci, (val, ct) in enumerate(zip(sample, col_types), 1):
        c = ws.cell(row=4, column=ci, value=val)
        c.font = nfont; c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if ct == "图片": c.fill = imfill
        elif ct == "AI": c.fill = aifill
        elif ct == "固定": c.fill = fxfill
        else: c.fill = hufill

    if len(sample) != ncols or len(col_types) != ncols:
        print(f"  ERROR {cat}: sample={len(sample)} col_types={len(col_types)} ncols={ncols}")
        continue

    for ri in range(5, 25):
        for ci in range(1, ncols + 1):
            c = ws.cell(row=ri, column=ci, value="")
            c.font = nfont; c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ct = col_types[ci - 1]
            if ct == "图片": c.fill = imfill
            elif ct == "AI": c.fill = aifill
            elif ct == "固定": c.fill = fxfill
            else: c.fill = PatternFill()

    # 尺码大小列填固定值
    for ci in range(1, ncols + 1):
        h = headers[ci - 1].split('\n')[0].strip()
        if h.startswith('尺码大小'):
            for ri in range(5, 25):
                c = ws.cell(row=ri, column=ci, value="S:80-95斤 M:95-105斤 L:105-115斤 XL:115-125斤")
                c.font = nfont; c.fill = fxfill; c.border = border
                c.alignment = Alignment(vertical="center")

    # 下拉验证 - 写入右侧隐藏列
    opt_start_col = ncols + 2
    field_ranges = {}

    for fi in fields:
        opts_str = fi['options']
        # 如果可选值为空但固定值含/分隔符，说明用户把选项写到了固定值列
        if not opts_str and fi['fixed'] and '/' in fi['fixed']:
            opts_str = fi['fixed']
            fi['options'] = opts_str  # 同步回去
        if opts_str:
            opts_list = [o.strip() for o in opts_str.split('/') if o.strip()]
            if opts_list:
                for oi, opt in enumerate(opts_list[:80]):
                    ws.cell(row=oi + 1, column=opt_start_col, value=opt)
                end_row = len(opts_list[:80])
                field_ranges[fi['field']] = (opt_start_col, 1, end_row)
                col_l = openpyxl.utils.get_column_letter(opt_start_col)
                ws.column_dimensions[col_l].hidden = True
                opt_start_col += 1

    # 给AI属性列加下拉 (跳过标题和短标题)
    for ci in range(1, ncols + 1):
        ct = col_types[ci - 1]
        if ct != "AI":
            continue
        fn = headers[ci - 1].split('\n')[0].replace('*', '').strip()
        # 标题和短标题不需要下拉
        if fn in ['商品标题', '短标题']:
            continue
        if fn in field_ranges:
            oc, sr, er = field_ranges[fn]
            col_l = openpyxl.utils.get_column_letter(oc)
            range_str = f"${col_l}${sr}:${col_l}${er}"
            dv = DataValidation(type="list", formula1=range_str, allow_blank=True)
            dv.error = "请从下拉选项中选择"
            dv.errorTitle = "无效输入"
            cell_col = openpyxl.utils.get_column_letter(ci)
            dv.add(f"{cell_col}5:{cell_col}200")
            ws.add_data_validation(dv)

    # 列宽 (根据内容类型)
    for ci in range(1, ncols + 1):
        cl = openpyxl.utils.get_column_letter(ci)
        h = headers[ci - 1].split('\n')[0].strip()
        ct = col_types[ci - 1]
        if ct == "图片":
            ws.column_dimensions[cl].width = 14
        elif '标题' in h and ct == 'AI':
            ws.column_dimensions[cl].width = 34  # 商品标题
        elif '短标题' in h:
            ws.column_dimensions[cl].width = 18
        elif '款号' in h:
            ws.column_dimensions[cl].width = 12
        elif '售价' in h:
            ws.column_dimensions[cl].width = 8
        elif 'SKU' in h:
            ws.column_dimensions[cl].width = 10
        elif '备注' in h:
            ws.column_dimensions[cl].width = 16
        else:
            ws.column_dimensions[cl].width = 13
    ws.freeze_panes = "A5"

    # 底部说明
    nr = 27
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=ncols)
    c = ws.cell(row=nr, column=1,
        value=f"SKU说明: 一行一个商品，多个颜色填SKU1~SKU6列。图片: 参考图->商品图/参考图/[文件名] | 主图1x1->商品图/主图1x1/[款号].jpg | 主图3x4->商品图/主图3x4/[款号].jpg")
    c.font = Font(name="微软雅黑", size=9, color="999999", italic=True)

# 保存
out = r"C:\Users\86153\Desktop\抖店上架助手\商品上架模板_v2.xlsx"
wb.save(out)
print(f"\n已生成: 商品上架模板_v2.xlsx")

# ===== 4. 保存JSON =====
config_json = {}
for cat in cat_order:
    config_json[cat] = {fi['field']: {'options': fi['options'], 'fixed': fi['fixed']}
                        for fi in categories[cat]}
with open(r'C:\Users\86153\Desktop\抖店上架助手\下拉选项配置.json', 'w', encoding='utf-8') as f:
    json.dump(config_json, f, ensure_ascii=False, indent=2)
print("已保存: 下拉选项配置.json")
print("\n完成!")
